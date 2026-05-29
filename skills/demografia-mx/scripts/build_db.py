#!/usr/bin/env python3
"""
demografia-mx: construye demografia.duckdb a partir de los XLSX CONAPO
(Conciliación Demográfica + Proyecciones 1950-2070).

Uso:
  python3 build_db.py [--db PATH]

Resultado:
  research/demografia.duckdb con tablas:
    - poblacion          año × entidad × edad × sexo × población
    - esperanza_vida     año × entidad × esperanza al nacer
    - indicadores        año × entidad × indicador × valor
"""
import argparse
import os
from pathlib import Path

import duckdb
import openpyxl

DEFAULT_BASE = Path("/Volumes/Storage/sistemia-skills-mx/research")
DEFAULT_DB = DEFAULT_BASE / "demografia.duckdb"
CONAPO_DIR = DEFAULT_BASE / "conapo/extracted/ConDem50a19_ProyPob20a70"


def stream_rows(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    header = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = [str(c).strip() if c is not None else "" for c in row]
            continue
        yield header, row


def load_poblacion(con):
    print("→ Cargando población (Pob_Mitad 1950-2070)…")
    xlsx = CONAPO_DIR / "0_Pob_Mitad_1950_2070.xlsx"
    con.execute("DROP TABLE IF EXISTS poblacion")
    con.execute("""
        CREATE TABLE poblacion (
            anio INTEGER,
            entidad VARCHAR,
            cve_geo INTEGER,
            edad INTEGER,
            sexo VARCHAR,
            poblacion BIGINT
        )
    """)
    rows = []
    n = 0
    for header, row in stream_rows(xlsx):
        try:
            rows.append((int(row[1]), row[2], int(row[3]), int(row[4]), row[5], int(row[6])))
        except (TypeError, ValueError):
            continue
        n += 1
        if len(rows) >= 50000:
            con.executemany("INSERT INTO poblacion VALUES (?, ?, ?, ?, ?, ?)", rows)
            rows = []
    if rows:
        con.executemany("INSERT INTO poblacion VALUES (?, ?, ?, ?, ?, ?)", rows)
    con.execute("CREATE INDEX IF NOT EXISTS idx_pob_anio ON poblacion(anio)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_pob_entidad ON poblacion(entidad)")
    print(f"  · {n:,} renglones cargados")


def load_esperanza(con):
    print("→ Cargando esperanza de vida…")
    xlsx = CONAPO_DIR / "6_Esperanza_Vida_Nacer_1950_2070.xlsx"
    if not xlsx.exists():
        print("  · no encontrado, saltando")
        return
    con.execute("DROP TABLE IF EXISTS esperanza_vida")
    con.execute("""
        CREATE TABLE esperanza_vida (
            anio INTEGER,
            entidad VARCHAR,
            sexo VARCHAR,
            esperanza DOUBLE
        )
    """)
    n = 0
    rows = []
    for header, row in stream_rows(xlsx):
        try:
            anio = int(row[1])
            entidad = row[2]
            sexo = row[4] if len(row) > 4 else "Total"
            esperanza = float(row[-1])
            rows.append((anio, entidad, sexo, esperanza))
            n += 1
        except (TypeError, ValueError):
            continue
        if len(rows) >= 50000:
            con.executemany("INSERT INTO esperanza_vida VALUES (?, ?, ?, ?)", rows)
            rows = []
    if rows:
        con.executemany("INSERT INTO esperanza_vida VALUES (?, ?, ?, ?)", rows)
    print(f"  · {n:,} renglones")


def add_vistas_utiles(con):
    print("→ Creando vistas útiles…")
    con.execute("""
        CREATE OR REPLACE VIEW poblacion_total AS
        SELECT anio, entidad, SUM(poblacion) AS poblacion
        FROM poblacion
        GROUP BY anio, entidad
    """)
    con.execute("""
        CREATE OR REPLACE VIEW poblacion_por_grupo AS
        SELECT
            anio, entidad, sexo,
            CASE
                WHEN edad < 5   THEN '0-4'
                WHEN edad < 15  THEN '5-14'
                WHEN edad < 18  THEN '15-17'
                WHEN edad < 25  THEN '18-24'
                WHEN edad < 35  THEN '25-34'
                WHEN edad < 45  THEN '35-44'
                WHEN edad < 55  THEN '45-54'
                WHEN edad < 65  THEN '55-64'
                ELSE '65+'
            END AS grupo_edad,
            SUM(poblacion) AS poblacion
        FROM poblacion
        GROUP BY anio, entidad, sexo, grupo_edad
    """)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", default=str(DEFAULT_DB))
    args = ap.parse_args()

    if not CONAPO_DIR.exists():
        raise SystemExit(f"ERROR: no encuentro {CONAPO_DIR}. Descarga el ZIP CONAPO Conciliación primero.")

    print(f"Construyendo {args.db}…")
    con = duckdb.connect(args.db)
    load_poblacion(con)
    load_esperanza(con)
    add_vistas_utiles(con)
    print("\nResumen:")
    print(con.execute("SELECT COUNT(*) AS rows FROM poblacion").fetchone())
    print(con.execute("SELECT MIN(anio), MAX(anio), COUNT(DISTINCT entidad) FROM poblacion").fetchone())
    print("\n✓ Listo")


if __name__ == "__main__":
    main()
